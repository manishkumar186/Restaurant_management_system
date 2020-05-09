from tkinter import *
from tkinter.ttk import Combobox, Treeview
from tkinter.messagebox import showinfo
import openpyxl
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from http.client import *
from connectionfile import *

class cart:


    def __init__(self):
        self.bid = 0
        self.listcart = []
        self.serialno = 1

        self.window = Toplevel()
        self.window.geometry("1010x630")
        self.window.resizable(FALSE, FALSE)

        self.frame1 = Frame(self.window, width=1010, height=60, bg="red", bd=8, relief="ridge")
        self.frame2 = Frame(self.window, width=1010, height=340, bg="orange", bd=10, relief="ridge")

        self.frame1.place(x=0, y=0)
        self.frame2.place(x=0, y=290)

        t = datetime.now()
        self.lst = str(t).split(" ")
        self.currentdate = self.lst[0]
        self.s = self.lst[1]
        self.lst3 = str(self.s).split(":")
        self.currenttime = self.lst3[0] + ":" + self.lst3[1]
        txt = "DATE=" + self.currentdate + "\n""TIME=" + self.currenttime
        self.lb4 = Label(self.window, text=str(txt))
        self.lb4.place(x=900, y=15)
        self.lb1 = Label(self.window, text="SELECT MENU", bg="red")
        self.lb2 = Label(self.window, text="SELECT QUANTITY", bg="red")
        self.lb5 = Label(self.window, text="PAYMENT TYPE", bg="orange")
        self.lb6 = Label(self.window, text="CASH COLLECTED", bg="orange")
        self.lb7 = Label(self.window, text="CASH RETURN", bg="orange")
        self.lb8 = Label(self.window, text="BILL TYPE", bg="orange")
        self.lb9 = Label(self.window, text="ADDRESS", bg="orange")
        self.lb10 = Label(self.window, text="MOBILE NO", bg="orange")
        self.lb5.place(x=10, y=300)
        self.lb6.place(x=10, y=350)
        self.lb7.place(x=10, y=400)
        self.lb8.place(x=10, y=450)
        self.lb9.place(x=10, y=500)
        self.lb10.place(x=10, y=550)
        self.cb3 = Combobox(self.window, values=("CASH", "CREDIT CARD", "PAYTM"), state="readonly")
        self.cb3.place(x=200, y=300)
        self.textbox1 = Entry(self.window)
        self.textbox2 = Entry(self.window)
        self.textbox3 = Entry(self.window)
        self.textbox4 = Entry(self.window)
        self.textbox1.place(x=200, y=350)
        self.textbox2.place(x=200, y=400)
        self.bt3 = Button(self.window, text="GENERATE BILL",bg="gold", command=self.insertintodb)
        self.bt3.place(x=100, y=580)

        self.bt4 = Button(self.window, text="CLOSE",bg="gold", command=self.window.destroy)
        self.bt4.place(x=600, y=580)

        self.textbox3.place(x=200, y=500)
        self.textbox4.place(x=200, y=550)
        self.cb4 = Combobox(self.window, values=("TAKE AWAY", "HOME DELIVERY", "DINEIN"), state="readonly")
        self.cb4.place(x=200, y=450)
        self.cb1 = Combobox(self.window, state="readonly")
        self.cb2 = Combobox(self.window, values=("1", "2", "3", "4", "5", "6", "7", "8", "9", "10"), state="readonly")
        self.bt1 = Button(self.window, text="ADD TO CART", command=self.add)
        self.totalbill = StringVar()
        self.lbtotal = Label(self.window, textvariable=self.totalbill)
        self.treeview = Treeview(self.window,
                                 columns=("SERIAL NUMBER", "ITEM NAME", "PRICE", "QUANTITY", "TOTAL PRIZE"))
        self.treeview.heading("SERIAL NUMBER", text="SERIAL NUMBER")
        self.treeview.heading("ITEM NAME", text="ITEM NAME")
        self.treeview.heading("PRICE", text="PRICE")
        self.treeview.heading("QUANTITY", text="QUANTITY")
        self.treeview.heading("TOTAL PRIZE", text="TOTAL PRIZE")
        self.treeview["show"] = "headings"
        self.lb1.place(x=30, y=15)
        self.lb2.place(x=450, y=15)
        self.cb1.place(x=150, y=15)
        self.cb2.place(x=580, y=15)
        self.bt1.place(x=750, y=15)
        self.treeview.place(x=0, y=60)
        self.lbtotal.place(x=500, y=300)
        self.totalbill.set("Total Bill is: Rs. 0")
        l = []
        conn = connectionfile.connect('')
        s = "select menuname from hotel"
        cr = conn.cursor()
        cr.execute(s)
        result = cr.fetchall()
        for row in result:
            menuid = row[0]
            l.append(str(menuid))
        self.cb1["values"] = l

        self.window.mainloop()

    def add(self):
        tp = 0.0
        itemname = self.cb1.get()
        quantity = self.cb2.get()
        if itemname == "":
            showinfo("", "choose item name")
        elif quantity == "":
            showinfo('', 'choose quantity')
        else:

            flag = FALSE
            matchingindex = 0
            sl = []
            for i in range(0, len(self.listcart)):
                ln = self.listcart[i]
                if ln[1] == itemname:
                    matchingindex = i
                    flag = TRUE
                    sl = ln
                    break
            if flag == FALSE:
                smallist = []
                smallist.append(self.serialno)
                self.serialno = self.serialno + 1
                smallist.append(itemname)
                conn = connectionfile.connect('')
                s = "select price from hotel where menuname='" + self.cb1.get() + "'"
                cr = conn.cursor()
                cr.execute(s)
                result = cr.fetchone()
                smallist.append(str(result[0]))
                smallist.append(quantity)
                total = (float)(result[0]) * (int)(quantity)
                smallist.append(total)
                self.listcart.append(smallist)
                i = 0
                for row in self.treeview.get_children():
                    self.treeview.delete(row)
                for row in self.listcart:
                    tp = tp + (float)(row[4])
                    self.treeview.insert("", index=i, values=row)
                    i = i + 1
                netamount = 0.0
                gst = (5 * tp) / 100
                netamount = tp + gst
                self.totalbill.set(
                    "Total Bill is: Rs. " + str(tp) + "\n  GST 5%" + "\n NetAmount Rs. " + str(netamount))
            else:
                oldquantity = (int)(sl[3])
                newquantity = (int)(quantity)
                finalquantity = oldquantity + newquantity
                conn = connectionfile.connect('')
                s = "select price from hotel where menuname='" + self.cb1.get() + "'"
                cr = conn.cursor()
                cr.execute(s)
                result = cr.fetchone()
                total = (float)(result[0]) * (int)(finalquantity)
                sl[3] = finalquantity
                sl[4] = total
                self.listcart[matchingindex] = sl
                i = 0
                for row in self.treeview.get_children():
                    self.treeview.delete(row)
                for row in self.listcart:
                    tp = tp + (float)(row[4])
                    self.treeview.insert("", index=i, values=row)
                    i = i + 1
                netamount = 0.0
                gst = (5 * tp) / 100
                netamount = tp + gst
                self.totalbill.set(
                    "Total Bill is: Rs. " + str(tp) + "\n  GST 5%" + "\n NetAmount Rs. " + str(netamount))

    def insertintodb(self):
        # print()

        paymentmethod = self.cb3.get()
        billtype = self.cb4.get()
        if paymentmethod == "":
            showinfo('', 'SELECT PAYMENT METHODE')
        elif billtype == "":
            showinfo('', 'SELECT BILL TYPE')
        else:
            if paymentmethod == "CASH" and self.textbox1.get() == "":
                showinfo('', 'ENTER COLLECTED CASH')



            elif billtype == "HOME DELIVERY" and self.textbox3.get() == "":
                showinfo('', 'ENTER ADDRESS')
            elif self.textbox4.get() == "":
                showinfo('', 'ENTER MOBILE NUMBER')
            elif(self.textbox4.get().isalpha()==True):
                showinfo("", "PLEASE ENTER VALID MOBILE NUMBER")
            else:
                if paymentmethod != "CASH":
                    self.textbox1.insert(0, 0.0)
                    self.textbox2.insert(0, 0.0)
                if billtype != "HOME DELIVERY":
                    self.textbox3.insert(0, "")

                conn = connectionfile.connect('')
                cr = conn.cursor()
                bill = 0.0
                for row in self.listcart:
                    bill = bill + (float)(row[4])
                netamount = 0.0
                gst = (5 * bill) / 100
                netamount = bill + gst
                query1 = "insert into bill values(NULL,'" + str(
                    self.currentdate) + "','" + self.textbox4.get() + "','" + self.cb4.get() + "','" + self.textbox3.get() + "','" + str(
                    netamount) + "','" + str(
                    5.0) + "','" + self.cb3.get() + "','" + self.textbox1.get() + "','" + self.textbox2.get() + "','pending')"
                cr.execute(query1)
                conn.commit()
                query2 = "select billid from bill"
                cr.execute(query2)
                rs = cr.fetchall()
                billid = 0
                for row in rs:
                    billid = row[0]
                self.bid = billid
                for row in self.listcart:
                    query3 = "select order_number from hotel where menuname='" + str(row[1]) + "'"
                    cr.execute(query3)
                    rs = cr.fetchone()
                    menuid = str(rs[0])
                    query4 = "insert into billdetail values(NULL,'" + menuid + "','" + str(row[2]) + "','" + str(
                        row[3]) + "','" + str(row[4]) + "','" + str(billid) + "')"
                    cr.execute(query4)
                    conn.commit()
                self.downloadbill()
                msg = "ORDER SUCCESSFULLY"
                mobile = self.textbox4.get()
                msg = msg.replace(" ", "%20")
                conn = HTTPConnection("server1.vmm.education")
                conn.request('GET',
                             "/VMMCloudMessaging/AWS_SMS_Sender?""username=manishkumar&password=A0809UGZ&message=" + msg + "&phone_numbers=" + mobile)
                response = conn.getresponse()
                # print(response.read())


    def downloadbill(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.row_dimensions[1].height = 25
        sheet.row_dimensions[2].height = 25
        sheet.row_dimensions[3].height = 25
        sheet.row_dimensions[4].height = 25
        sheet.row_dimensions[5].height = 25

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
        index = 1
        sheet.cell(row=index, column=1).value = 'SERIAL NUMBER'
        sheet.cell(row=index, column=1).font = Font(size=18, bold=True)
        sheet.cell(row=index, column=2).value = 'MENU ITEM'
        sheet.cell(row=index, column=2).font = Font(size=18, bold=True)
        sheet.cell(row=index, column=3).value = 'PRICE'
        sheet.cell(row=index, column=3).font = Font(size=18, bold=True)
        sheet.cell(row=index, column=4).value = 'QUANTITY'
        sheet.cell(row=index, column=4).font = Font(size=18, bold=True)
        sheet.cell(row=index, column=5).value = 'TOTAL'
        sheet.cell(row=index, column=5).font = Font(size=18, bold=True)
        index = index + 1
        for row in self.treeview.get_children():
            d = self.treeview.item(row)
            l = d['values']

            sheet.cell(row=index, column=1).value = l[0]
            sheet.cell(row=index, column=1).font = Font(size=16, name='Times New Roman')
            sheet.cell(row=index, column=1).alignment = Alignment(horizontal='center')
            sheet.cell(row=index, column=2).value = l[1]
            sheet.cell(row=index, column=2).font = Font(size=16, name='Times New Roman')
            sheet.cell(row=index, column=2).alignment = Alignment(horizontal='center')
            sheet.cell(row=index, column=3).value = l[2]
            sheet.cell(row=index, column=3).font = Font(size=16, name='Times New Roman')
            sheet.cell(row=index, column=3).alignment = Alignment(horizontal='center')
            sheet.cell(row=index, column=4).value = l[3]
            sheet.cell(row=index, column=4).font = Font(size=16, name='Times New Roman')
            sheet.cell(row=index, column=4).alignment = Alignment(horizontal='center')
            sheet.cell(row=index, column=5).value = l[4]
            sheet.cell(row=index, column=5).font = Font(size=16, name='Times New Roman')
            sheet.cell(row=index, column=5).alignment = Alignment(horizontal='center')
            index = index + 1
        index = index + 2
        bill = 0.0
        for row in self.listcart:
            bill = bill + (float)(row[4])
        netamount = 0.0
        gst = (5 * bill) / 100
        netamount = bill + gst
        sheet.cell(row=index, column=1).value = 'TOTAL PRICE '
        sheet.cell(row=index, column=1).font = Font(size=15, bold=True)
        sheet.cell(row=index, column=2).value = 'Rs. ' + str(bill)
        sheet.cell(row=index, column=2).font = Font(size=15, bold=True)
        index = index + 1
        sheet.cell(row=index, column=1).value = 'GST '
        sheet.cell(row=index, column=1).font = Font(size=15, bold=True)
        sheet.cell(row=index, column=2).value = '5%'
        sheet.cell(row=index, column=2).font = Font(size=15, bold=True)
        index = index + 1
        sheet.cell(row=index, column=1).value = 'NET  AMOUNT '
        sheet.cell(row=index, column=1).font = Font(size=15, bold=True)
        sheet.cell(row=index, column=2).value = 'Rs. ' + str(netamount)
        sheet.cell(row=index, column=2).font = Font(size=15, bold=True)

        showinfo('', 'Bill Excel Sheet Generated Successfully \n storage path \n ' + "bills/billno_" + str(
            self.bid) + '.xlsx')
        wb.save("bills/billno_" + str(self.bid) + '.xlsx')

#obj=cart()







